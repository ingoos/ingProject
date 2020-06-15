# 잉구 맥북 
import time
start = time.time()
from openpyxl import load_workbook
from openpyxl import Workbook
from urllib.request import urlopen
from urllib.parse import quote_plus
from bs4 import BeautifulSoup
import ssl

rank = []

wb = load_workbook('test.xlsx',read_only=True)
data = wb.active

data_range = data['A2':'B9']
context = ssl._create_unverified_context()

for item in data_range:
    rank_tmp = []
    keyword = item[0].value
    encode = quote_plus(keyword)
    url = 'https://search.naver.com/search.naver?sm=top_hty&fbm=0&ie=utf8&query='+encode
    html = urlopen(url,context=context)
    obj = BeautifulSoup(html,"html.parser")

    #연관검색 텍스트 가져오기
    incard = obj.select('#nx_related_keywords > dl > dd.lst_relate._related_keyword_list > ul > li > a')
    for card in incard:
        rank_tmp.append(card.text)

    rank.append(rank_tmp)
    rank_tmp = []
wb2 = Workbook()
ws = wb2.active

for item in rank:
    ws.append(item)

wb2.save('test2.xlsx')

print(rank)
print(time.time()-start)